#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from DayFeatureEx import DayFeatureEx
from datetime import timedelta


def WriteDataSet(f_path, data_set):
    wb = openpyxl.Workbook()
    ws = wb.active

    header = [u"№", "PtId", "DT_Fix", "NoctMin_T", "NoctMin_Gl", "Ill_months", "Age", "Gender", "Height", "Weight", "InsMod"]
    max_rises_count = max(map(lambda x: len(x.GetGlRises()), data_set))
    for i in range(1, max_rises_count+1):
        header.append("BMax_T" + str(i))
        header.append("BMax_Gl" + str(i))
        header.append("Max_T" + str(i))
        header.append("Max_Gl" + str(i))

    for col, col_name in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=col_name)

    for index, record in enumerate(data_set, start=1):
        ws.cell(row=index+1, column=1, value=index)
        ws.cell(row=index+1, column=2, value=record.GetPtId())
        ws.cell(row=index+1, column=3, value=record.GetFixedDT().strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=index+1, column=4, value=record.GetNocturnalMinimum()[0].total_seconds())
        ws.cell(row=index+1, column=5, value=record.GetNocturnalMinimum()[1])
        ws.cell(row=index+1, column=6, value=record.GetIllMonths())
        ws.cell(row=index+1, column=7, value=record.GetAge())
        ws.cell(row=index+1, column=8, value=record.GetGender())
        ws.cell(row=index+1, column=9, value=record.GetHeight())
        ws.cell(row=index+1, column=10, value=record.GetWeight())
        ws.cell(row=index+1, column=11, value=record.GetInsMod())

        gl_rises = record.GetGlRises()
        for r_count in range(0, max_rises_count):
            if r_count < len(gl_rises):
                gl_rise = gl_rises[r_count]
                ws.cell(row=index+1, column=11 + r_count*4 + 1, value=gl_rise[0][0].total_seconds())
                ws.cell(row=index+1, column=11 + r_count*4 + 2, value=gl_rise[0][1])
                ws.cell(row=index+1, column=11 + r_count*4 + 3, value=gl_rise[1][0].total_seconds())
                ws.cell(row=index+1, column=11 + r_count*4 + 4, value=gl_rise[1][1])
            else:
                ws.cell(row=index+1, column=11 + r_count*4 + 1, value="NA")
                ws.cell(row=index+1, column=11 + r_count*4 + 2, value="NA")
                ws.cell(row=index+1, column=11 + r_count*4 + 3, value="NA")
                ws.cell(row=index+1, column=11 + r_count*4 + 4, value="NA")

    wb.save(f_path)
