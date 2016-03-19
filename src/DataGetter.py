import openpyxl


class PatientGlucoseMeasurement:
    def __init__(self, i_pt_id, i_datetime, i_glucose_level):
        self.m_pt_id = i_pt_id
        self.m_datetime = i_datetime
        self.m_glucose_level = i_glucose_level

    def __str__(self):
        return 'Pt_Id=' + str(self.GetPtId()) + ', Dt=' + str(self.GetDateTime()) + ', Gl=' + str(self.GetGlucoseLevel())

    def GetPtId(self):
        return self.m_pt_id

    def GetDateTime(self):
        return self.m_datetime

    def GetGlucoseLevel(self):
        return self.m_glucose_level


class DataGetter:
    def __init__(self,
                 i_work_book_name,
                 i_patient_id_column_num,
                 i_datetime_column_num,
                 i_glucose_level_column_num):
        wb = openpyxl.load_workbook(i_work_book_name, data_only=True)
        self.m_data_sheet = wb.active
        self.m_current_row = 2
        self.m_patient_id_column_num = i_patient_id_column_num
        self.m_datetime_column_num = i_datetime_column_num
        self.m_glucose_level_column_num = i_glucose_level_column_num

    def LookAtNextRow(self):
        patient_id = self.m_data_sheet.cell(row=self.m_current_row, column=self.m_patient_id_column_num).value
        patient_datetime = self.m_data_sheet.cell(row=self.m_current_row, column=self.m_datetime_column_num).value
        patient_glucose = self.m_data_sheet.cell(row=self.m_current_row, column=self.m_glucose_level_column_num).value
        return [patient_id, patient_datetime, patient_glucose]

    def IsDataProcessed(self):
        return self.LookAtNextRow()[0] is None

    def GetNextRow(self):
        result = self.LookAtNextRow()
        self.m_current_row += 1
        return result

    def GetNextMeasurement(self):
        next_row = self.GetNextRow()
        return PatientGlucoseMeasurement(next_row[0], next_row[1], next_row[2])

    def _GetAllMeasurements(self):
        measurements = []
        while not self.IsDataProcessed():
            measurements.append(self.GetNextMeasurement())
        return measurements

    def GetDataHeader(self):
        return ['PtID', 'DataTime', 'Gl']

    @staticmethod
    def GetAllMeasurements(i_work_book_name,
                           i_patient_id_column_num,
                           i_datetime_column_num,
                           i_glucose_level_column_num):
        data_getter = DataGetter(i_work_book_name,
                                 i_patient_id_column_num,
                                 i_datetime_column_num,
                                 i_glucose_level_column_num)
        return data_getter._GetAllMeasurements()



def CreateExternalDataGetter(i_external_work_book_name):
    patient_id_column_num = 2
    datetime_column_num = 6
    glucose_level_column_num = 5
    return DataGetter(i_external_work_book_name,
                      patient_id_column_num,
                      datetime_column_num,
                      glucose_level_column_num)


def CreateInternalDataGetter(i_internal_work_book_name):
    patient_id_column_num = 1
    datetime_column_num = 2
    glucose_level_column_num = 3
    return DataGetter(i_internal_work_book_name,
                      patient_id_column_num,
                      datetime_column_num,
                      glucose_level_column_num)