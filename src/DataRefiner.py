#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os

from DataGetter import DataGetter
from FeatureExtraction import FeatureExtractor
from FeatureExtraction import ExtremaFilter
from FeatureExtraction import SeriesModifier as SM
from Features import RiseFeature, DayFeature
import datetime

import matplotlib.pyplot as plt
import tkMessageBox
import tkFileDialog
import matplotlib.dates
from matplotlib.widgets import Button

import pickle
from openpyxl import Workbook

class DayMeasurementsIterator:
    g_ts = datetime.datetime.strptime("05:00:00", "%H:%M:%S").time()

    def __init__(self, root_path):
        self.m_files = list()
        if os.path.isfile(root_path):
            if root_path.endswith('.xlsx'):
                self.m_files.append(root_path)
        elif os.path.isdir(root_path):
            for root, dirs, files in os.walk(root_path):
                for cur_file in files:
                    if cur_file.endswith('.xlsx'):
                        src_path = os.path.join(root, cur_file)
                        self.m_files.append(src_path)

        self.m_next_file_index = 0
        self.m_cur_day_measurements = list()
        self.m_next_day_measurement_index = 0
        self.m_current = None

    def is_empty(self):
        is_measurements_empty = self.m_next_day_measurement_index == len(self.m_cur_day_measurements)
        is_files_empty = self.m_next_file_index == len(self.m_files)
        return is_measurements_empty and is_files_empty

    def _read_next_file(self):
        assert self.m_next_file_index < len(self.m_files)
        src_file = self.m_files[self.m_next_file_index]
        con_ms = DataGetter.GetAllMeasurements(src_file,
                                               i_patient_id_column_num=1,
                                               i_datetime_column_num=2,
                                               i_glucose_level_column_num=3)
        self.m_cur_day_measurements = FeatureExtractor.SeparateMeasurementsForDays(con_ms, DayMeasurementsIterator.g_ts)
        self.m_next_file_index += 1
        self.m_next_day_measurement_index = 0

    @staticmethod
    def _NormaliseMeasurements(measurements):
        without_nulls = SM.EraseNulls(measurements)
        return without_nulls

    def next(self):
        if self.is_empty():
            raise EOFError("there is no more measurements")
        if self.m_next_day_measurement_index == len(self.m_cur_day_measurements):
            self._read_next_file()
        measurements = self.m_cur_day_measurements[self.m_next_day_measurement_index]
        self.m_current = self._NormaliseMeasurements(measurements)
        self.m_next_day_measurement_index += 1
        return self.m_current

    def current(self):
        return self.m_current


class GUI:
    def __init__(self):
        axcolor = 'lightgoldenrodyellow'
        acceptax = plt.axes([0.4, 0.025, 0.1, 0.04])
        self.button_stop = Button(acceptax, 'Stop', color=axcolor, hovercolor='0.975')
        acceptax = plt.axes([0.6, 0.025, 0.1, 0.04])
        self.button_accept = Button(acceptax, 'Accept', color=axcolor, hovercolor='0.975')
        rejectax = plt.axes([0.8, 0.025, 0.1, 0.04])
        self.button_reject = Button(rejectax, 'Reject', color=axcolor, hovercolor='0.975')

        self.fig = plt.gcf()
        self.ax = self.fig.add_subplot(111)
        plt.subplots_adjust(bottom=0.25)
        self.status_text_id = None
        self.mh_preview_point_list = list()

        self.m_on_chart_l_clicked_action = None
        self.m_on_chart_r_clicked_action = None
        self.m_on_chart_m_clicked_action = None
        self.fig.canvas.mpl_connect('button_press_event', self._OnFigureClicked)

        self.m_on_undo_action = None
        self.fig.canvas.mpl_connect('key_press_event', self._OnKeyPressed)

    def _OnFigureClicked(self, event):
        if event.inaxes == self.ax:
            tmp_dt = matplotlib.dates.num2date(event.xdata)
            dt = datetime.datetime(
                year=tmp_dt.year,
                month=tmp_dt.month,
                day=tmp_dt.day,
                hour=tmp_dt.hour,
                minute=tmp_dt.minute,
                second=tmp_dt.second
            )
            print 'Pick: ' + str(dt) + ' == ' + str(event.ydata)
            if event.button == 1:
                self.m_on_chart_l_clicked_action(dt)
            elif event.button == 2:
                self.m_on_chart_m_clicked_action(dt)
            elif event.button == 3:
                self.m_on_chart_r_clicked_action(dt)

    def _OnKeyPressed(self, event):
        if event.key == 'ctrl+z':
            self.m_on_undo_action()

    def AskForInputData(self):
        root_path = tkFileDialog.askopenfilename(title='Select data to process',
                                                 filetypes=[('Excel files', '.xlsx')])
        return root_path

    def AskForFileToSave(self):
        file_path = tkFileDialog.asksaveasfilename(title='Save As..',
                                                   filetypes=[('Excel files', '.xlsx')])
        return file_path

    def ShowMessage(self, msg_show):
        tkMessageBox.showinfo("Info", msg_show)
        return

    def Show(self):
        plt.show()

    def PlotChart(self, title, x_data, y_data):
        self.ax.cla()
        self.mh_preview_point_list = list()
        self.ax.plot(x_data, y_data, 'b')
        plt.title(title)
        plt.draw()

    def UpdateStatusLine(self, text):
        if self.status_text_id:
            self.fig.texts.remove(self.status_text_id)
        self.status_text_id = self.fig.text(0, 0.02, text)

    def _AddPointPreview(self, x_data, y_data, style):
        h, = self.ax.plot(x_data, y_data, style)
        self.mh_preview_point_list.append(h)
        plt.draw()
        print 'Preview added'

    def AddBeforeMaxPreview(self, x_data, y_data):
        self._AddPointPreview(x_data, y_data, 'bo')

    def AddMaxPreview(self, x_data, y_data):
        self._AddPointPreview(x_data, y_data, 'ro')

    def AddNocturnalMinimumPreview(self, x_data, y_data):
        self._AddPointPreview(x_data, y_data, 'go')

    def DeleteLastAddedPreview(self):
        if self.mh_preview_point_list:
            h = self.mh_preview_point_list.pop()
            h.remove()
            plt.draw()
            print 'Preview deleted'

    def SetOnAcceptAction(self, on_accept_action):
        self.button_accept.on_clicked(lambda x: on_accept_action())

    def SetOnRejectAction(self, on_reject_action):
        self.button_reject.on_clicked(lambda x: on_reject_action())

    def SetOnStopAction(self, on_stop_action):
        self.button_stop.on_clicked(lambda x: on_stop_action())

    def SetOnUndoAction(self, on_undo_action):
        self.m_on_undo_action = on_undo_action

    def SetOnPickBeforeMaxAction(self, on_pick_before_max):
        self.m_on_chart_l_clicked_action = on_pick_before_max

    def SetOnPickMaxAction(self, on_pick_max):
        self.m_on_chart_m_clicked_action = on_pick_max

    def SetOnPickNocturnalMinimumAction(self, on_pick_nocturnal_minimum):
        self.m_on_chart_r_clicked_action = on_pick_nocturnal_minimum


class DataRefiner:
    def __init__(self):
        self.m_view = GUI()
        self.day_measurements_iterator = None
        self.m_picked_measurements = list()
        self.m_selected_day_features = list()
        self.processed_measurements = 0
        self.undo_stack = list()

        self._ConnectHandlers()

    def _ConnectHandlers(self):
        self.m_view.SetOnStopAction(self._OnStopAction)
        self.m_view.SetOnAcceptAction(self._OnAcceptAction)
        self.m_view.SetOnRejectAction(self._OnRejectAction)
        self.m_view.SetOnPickBeforeMaxAction(self._OnPickBeforeMaxAction)
        self.m_view.SetOnPickMaxAction(self._OnPickMaxAction)
        self.m_view.SetOnPickNocturnalMinimumAction(self._OnPickNocturnalMinimumAction)
        self.m_view.SetOnUndoAction(self._OnUndoAction)

    def _SerializeSelectedDayFeatures(self, file_name):
        wb = Workbook()
        ws = wb.active
        cur_row = 1
        cur_col = 1

        ws.cell(row=cur_row, column=cur_col, value='Record Index')
        ws.cell(row=cur_row, column=cur_col+1, value='PtId')
        cur_col += 2
        rf_count = max(map(lambda x: len(x.GetRiseFeatures()), self.m_selected_day_features))
        for rf_index in range(0, rf_count):
            ws.cell(row=cur_row, column=cur_col, value='BeforeMax_' + str(rf_index+1) + '_Dt')
            ws.cell(row=cur_row, column=cur_col+1, value='BeforeMax_' + str(rf_index+1) + '_Gl')
            ws.cell(row=cur_row, column=cur_col+2, value='Max_' + str(rf_index+1) + '_Dt')
            ws.cell(row=cur_row, column=cur_col+3, value='Max_' + str(rf_index+1) + '_Gl')
            cur_col += 4
        ws.cell(row=cur_row, column=cur_col, value='NocturnalMin_Dt')
        ws.cell(row=cur_row, column=cur_col+1, value='NocturnalMin_Gl')
        cur_row += 1
        cur_col = 1

        for ind, df in enumerate(self.m_selected_day_features, start=1):
            ws.cell(row=cur_row, column=cur_col, value=ind)
            ws.cell(row=cur_row, column=cur_col+1, value=df.GetNocturnalMinimum().GetPtId())
            cur_col += 2
            for rf in df.GetRiseFeatures():
                ws.cell(row=cur_row, column=cur_col, value=rf.GetBeforeMax().GetDateTime())
                ws.cell(row=cur_row, column=cur_col+1, value=rf.GetBeforeMax().GetGlucoseLevel())
                ws.cell(row=cur_row, column=cur_col+2, value=rf.GetMax().GetDateTime())
                ws.cell(row=cur_row, column=cur_col+3, value=rf.GetMax().GetGlucoseLevel())
                cur_col += 4
            for i in range(0, rf_count-len(df.GetRiseFeatures())):
                ws.cell(row=cur_row, column=cur_col, value='-')
                ws.cell(row=cur_row, column=cur_col+1, value='-')
                ws.cell(row=cur_row, column=cur_col+2, value='-')
                ws.cell(row=cur_row, column=cur_col+3, value='-')
                cur_col += 4
            ws.cell(row=cur_row, column=cur_col, value=df.GetNocturnalMinimum().GetDateTime())
            ws.cell(row=cur_row, column=cur_col+1, value=df.GetNocturnalMinimum().GetGlucoseLevel())
            cur_row += 1
            cur_col = 1

        wb.save(file_name)

    def _OnStopAction(self):
        file_name = self.m_view.AskForFileToSave()
        self._SerializeSelectedDayFeatures(file_name)
        self.m_view.ShowMessage(self._GetStatus())
        exit()

    def _IntroduceMeasurementsToUser(self, measurements):
        x_data = map(lambda m: m.GetDateTime(), measurements)
        y_data = map(lambda m: m.GetGlucoseLevel(), measurements)
        mm = measurements[0]
        title = 'Pt=' + str(mm.GetPtId()) + ', Day=' + str(mm.GetDateTime().date())
        self.m_view.PlotChart(title, x_data, y_data)

    def _IntroduceNextMeasurementsOrStop(self):
        self.m_picked_measurements = list()
        self.m_view.UpdateStatusLine(self._GetStatus())
        while not self.day_measurements_iterator.is_empty():
            measurements = self.day_measurements_iterator.next()
            if not measurements:
                continue
            self._IntroduceMeasurementsToUser(measurements)
            return
        self._OnStopAction()

    @staticmethod
    def _ComposeDayFeature(picked_measurements):
        rise_features = list()
        for i in range(0, (len(picked_measurements) - 1) / 2):
            rise_features.append(RiseFeature(i_before_max=picked_measurements[i*2][1], i_max=picked_measurements[i*2+1][1]))
        return DayFeature(i_rise_features=rise_features, i_nocturnal_minimum=picked_measurements[-1][1])

    @staticmethod
    def _IsValidPicking(picked_measurements):
        if not picked_measurements:
            return False

        expect_after = {
            'before-max': 'max',
            'max': ('before-max', 'nocturnal-minimum'),
            'nocturnal-minimum': ()
        }

        expected_types = 'before-max'
        for pick in picked_measurements:
            if pick[0] in expected_types:
                expected_types = expect_after[pick[0]]
            else:
                return False
        if expected_types != ():
            return False
        return True

    def _GetStatus(self):
        status = 'Measurements processed: ' + str(len(self.m_selected_day_features)) + '/' + str(self.processed_measurements)
        return status

    def _OnAcceptAction(self):
        if DataRefiner._IsValidPicking(self.m_picked_measurements):
            day_feature = DataRefiner._ComposeDayFeature(self.m_picked_measurements)
            if day_feature and day_feature.IsValid():
                self.m_selected_day_features.append(day_feature)
                self.processed_measurements += 1
                self._IntroduceNextMeasurementsOrStop()
                return
        self.m_view.ShowMessage('Inappropriate picking')

    def _OnRejectAction(self):
        self.processed_measurements += 1
        self._IntroduceNextMeasurementsOrStop()

    @staticmethod
    def _FindNearest(measurements, dt, alternative_return=None):
        nearest = alternative_return
        a, b = 0, len(measurements) - 1
        if b <= a:
            return nearest

        while b - a > 1:
            m = (a + b) / 2
            m_dt = measurements[m].GetDateTime()
            if dt < m_dt:
                b = m
            else:
                a = m

        nearest = min(measurements[a], measurements[b], key=lambda x: abs(x.GetDateTime() - dt))
        return nearest

    def _OnPickBeforeMaxAction(self, x_data):
        cur_ms = self.day_measurements_iterator.current()
        if not cur_ms:
            return

        nearest_m = DataRefiner._FindNearest(cur_ms, x_data)
        if not nearest_m:
            return

        # // not robust
        # minimas = ExtremaFilter.FindExtremalMeasurements(cur_ms, 'Min')
        # m = DataRefiner._FindNearest(minimas, x_data, nearest_m)
        m = nearest_m

        self.m_picked_measurements.append(('before-max', m))
        self.m_view.AddBeforeMaxPreview(m.GetDateTime(), m.GetGlucoseLevel())

        self.undo_stack.append(self._UndoPicking)

    def _OnPickMaxAction(self, x_data):
        cur_ms = self.day_measurements_iterator.current()
        if not cur_ms:
            return

        nearest_m = DataRefiner._FindNearest(cur_ms, x_data)
        if not nearest_m:
            return

        # // not robust
        # maximas = ExtremaFilter.FindExtremalMeasurements(cur_ms, 'Max')
        # m = DataRefiner._FindNearest(maximas, x_data, nearest_m)
        m = nearest_m

        self.m_picked_measurements.append(('max', m))
        self.m_view.AddMaxPreview(m.GetDateTime(), m.GetGlucoseLevel())

        self.undo_stack.append(self._UndoPicking)

    def _OnPickNocturnalMinimumAction(self, x_data):
        cur_ms = self.day_measurements_iterator.current()
        if not cur_ms:
            return

        next_day_date = cur_ms[0].GetDateTime().date() + datetime.timedelta(days=1)
        night_start = datetime.datetime.combine(next_day_date, datetime.time(hour=0))
        night_end = datetime.datetime.combine(next_day_date, datetime.time(hour=5))

        night_measurements = [x for x in cur_ms if night_start <= x.GetDateTime() <= night_end]
        if not night_measurements:
            return

        nocturnal_minimum = min(night_measurements, key=lambda m: m.GetGlucoseLevel())
        if not nocturnal_minimum:
            return

        self.m_picked_measurements.append(('nocturnal-minimum', nocturnal_minimum))
        self.m_view.AddNocturnalMinimumPreview(nocturnal_minimum.GetDateTime(), nocturnal_minimum.GetGlucoseLevel())

        self.undo_stack.append(self._UndoPicking)

    def _UndoPicking(self):
        self.m_view.DeleteLastAddedPreview()
        self.m_picked_measurements.pop()

    def _OnUndoAction(self):
        if self.undo_stack:
            f = self.undo_stack.pop()
            f()

    def Run(self):
        root_path = self.m_view.AskForInputData()
        self.day_measurements_iterator = DayMeasurementsIterator(root_path)
        self._IntroduceNextMeasurementsOrStop()
        self.m_view.Show()

refiner = DataRefiner()
refiner.Run()
